import openpyxl
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# mes cabecalho
def carregar_pagina(mes_arquivo_diretorio, nome_pagina):
    planilha_mes_1 = openpyxl.load_workbook(mes_arquivo_diretorio, data_only=True)
    pagina_mes = planilha_mes_1[nome_pagina]
    return pagina_mes


# mes_puxar_informacoes
from lib.mes_custo_mensal import get_custo_mensal, get_mes_atual, func_deu_erro
# from mes_inventario import get_inventario
from lib.mes_mapa_de_vendas import get_mapa_de_vendas
# from mes_producao_mensal import get_producao_mensal


# ano cabeçalho
caminho_arquivo_ano = "arquivos/dre_ano/DRE_ANO.xlsx"
planilha_ano = openpyxl.load_workbook(caminho_arquivo_ano, data_only=True)
planilha_ano_write = openpyxl.load_workbook(caminho_arquivo_ano)


def carregar_pagina_ano(nome_pagina):
    pagina = planilha_ano[nome_pagina]
    return pagina


def carregar_pagina_ano_write(nome_pagina):
    pagina = planilha_ano_write[nome_pagina]
    return pagina


# ano_main
pagina = carregar_pagina_ano("DRE")
pagina_write = carregar_pagina_ano_write("DRE")


def puxar_coluna_meses():
    coluna_meses_dre_ano = {}
    for indice, valor in enumerate(
        pagina.iter_cols(min_row=2, max_row=2, min_col=2, max_col=49, values_only=True),
        start=2,
    ):
        if valor[0]:
            coluna_meses_dre_ano[valor[0].lower()] = (
                indice + 1
            )  # + 1 para pegar valor de 'REAL'

    return coluna_meses_dre_ano


def puxar_linha_produto(tabela):
    linha_meses_dre_ano = {}
    pular_linhas = [
        "FATURAMENTO BRUTO",
        "IMPOSTOS SOB RECEITAS ( - )",
        "RECEITA LIQUIDA ( = )",
        "CMP (Custo do Material Produzido) ( - )",
        "CUSTO INDIRETO ( - )",
        "RESULTADO OPERACIONAL BRUTO ( = )",
        "DESPESAS COMERCIAIS ( - )",
        "DESP. ADMINISTRATIVAS ( - )",
        "DESP. C/ PESSOAL ( - )",
        "TOTAL DESPESAS OPERACIONAIS ( - )",
        "RESULTADO ANTES DO IR ( = )",
        "RESULTADO OPERACIONAL ( = )",
        "DESPESAS NÃO OPERACIONAIS ( - )",
        "Provisão Custo Férias e Décimo Terceiro",
    ]

    for indice, linha in enumerate(
        pagina.iter_rows(min_row=4, values_only=True), start=4
    ):
        nome = linha[0]
        if nome in pular_linhas:
            continue
        if nome and tabela.get(nome):
            linha_meses_dre_ano[nome] = indice

    return linha_meses_dre_ano


def substituir_valores_mapa_de_vendas(mes_arquivo_diretorio):
    mapa_de_vendas = get_mapa_de_vendas(mes_arquivo_diretorio)
    if func_deu_erro(): return
    # get_custo_mensal(mes_arquivo_diretorio) # Puxar mes atual
    for produto in mapa_de_vendas:
        if (
            produto
            and mapa_de_vendas.get(produto)
            and puxar_linha_produto(mapa_de_vendas).get(produto)
        ):
            linha_produto_atual = puxar_linha_produto(mapa_de_vendas)[produto]
            valor_atualizado = mapa_de_vendas[produto][2]
            pagina_write.cell(
                row=linha_produto_atual, column=puxar_coluna_meses()[get_mes_atual()]
            ).value = valor_atualizado


def substituir_valores_custo_mensal(mes_arquivo_diretorio):
    custo_mensal = get_custo_mensal(mes_arquivo_diretorio)
    if func_deu_erro(): return
    for produto in custo_mensal:
        if (
            produto
            and custo_mensal.get(produto)
            and puxar_linha_produto(custo_mensal).get(produto)
        ):
            linha_produto_atual = puxar_linha_produto(custo_mensal)[produto]
            valor_atualizado = (custo_mensal[produto][0] or 0) + (
                custo_mensal[produto][1] or 0
            )
            pagina_write.cell(
                row=linha_produto_atual, column=puxar_coluna_meses()[get_mes_atual()]
            ).value = valor_atualizado
